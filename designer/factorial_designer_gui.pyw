#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Factorial Designer GUI
DoE Tool
Add factors and levels
Combination counter shows the number of conditions

Milton F. Villegas - v0.9.4
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import itertools
import csv
import os
import re
from typing import Dict, List, Tuple, Optional

# Optional XLSX export
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

# Available factors with display names
AVAILABLE_FACTORS = {
    "buffer pH": "Buffer pH",
    "buffer_concentration": "Buffer Conc (mM)",
    "glycerol": "Glycerol (%)",
    "salt": "Salt (mM)",
    "dmso": "DMSO (%)",
    "detergent": "Detergent (%)",
}

def validate_numeric_input(action, char, entry_value):
    """Validate numeric input - allows digits, decimal, minus, comma for multiple values"""
    # Allow all deletions
    if action == '0':
        return True
    # Allow empty
    if char == '':
        return True
    # Allow digits, decimal point, negative sign, and comma for multiple entries
    return char.isdigit() or char in '.-,'

def validate_single_numeric_input(action, char, entry_value):
    """Single numeric value only (no commas)"""

    if action == '0':
        return True

    if char == '':
        return True

    return char.isdigit() or char in '.-'

class FactorModel:
    """Model with validation for factors and stock concentrations"""
    def __init__(self):
        self._factors: Dict[str, List[str]] = {}
        self._stock_concs: Dict[str, float] = {}
        
    def add_factor(self, name: str, levels: List[str], stock_conc: Optional[float] = None):
        name = name.strip()
        if not name:
            raise ValueError("Factor name cannot be empty.")
        if not levels:
            raise ValueError("At least one level is required.")
        self._factors[name] = list(levels)
        if stock_conc is not None:
            self._stock_concs[name] = stock_conc
    
    def update_factor(self, name: str, levels: List[str], stock_conc: Optional[float] = None):
        if name not in self._factors:
            raise ValueError(f"Factor '{name}' does not exist.")
        if not levels:
            raise ValueError("At least one level is required.")
        self._factors[name] = list(levels)
        if stock_conc is not None:
            self._stock_concs[name] = stock_conc
    
    def remove_factor(self, name: str):
        if name in self._factors:
            del self._factors[name]
        if name in self._stock_concs:
            del self._stock_concs[name]
    
    def get_factors(self) -> Dict[str, List[str]]:
        return {k: list(v) for k, v in self._factors.items()}
    
    def get_stock_conc(self, name: str) -> Optional[float]:
        return self._stock_concs.get(name)
    
    def get_all_stock_concs(self) -> Dict[str, float]:
        return dict(self._stock_concs)
    
    def clear(self):
        self._factors.clear()
        self._stock_concs.clear()
    
    def total_combinations(self) -> int:
        if not self._factors:
            return 0
        # Multiply all factor level counts (full factorial)
        result = 1
        for levels in self._factors.values():
            result *= len(levels)
        return result

class FactorEditDialog(tk.Toplevel):
    """Inline factor editor dialog with mandatory stock concentration"""
    def __init__(self, parent, factor_name: str, existing_levels: List[str] = None, 
                 stock_conc: Optional[float] = None):
        super().__init__(parent)

        display_name = AVAILABLE_FACTORS.get(factor_name, factor_name)
        
        self.title(f"Edit Factor: {display_name}")
        self.geometry("550x600")
        self.transient(parent)
        self.grab_set()
        
        self.factor_name = factor_name
        self.result_levels = None
        self.result_stock = None
        
        # Store entry widgets for key binding
        self.stock_entry_widget = None
        self.level_entry_widget = None
        
        # Validation commands with action code for proper backspace handling
        vcmd = (self.register(validate_numeric_input), '%d', '%S', '%P')  # For levels (allows commas)
        vcmd_stock = (self.register(validate_single_numeric_input), '%d', '%S', '%P')  # For stock (no commas)
        
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Factor info
        info_frame = ttk.LabelFrame(main_frame, text="Factor Information", padding=8)
        info_frame.pack(fill=tk.X, pady=(0, 8))
        
        ttk.Label(info_frame, text=f"Factor: {display_name}", 
                 font=("TkDefaultFont", 10, "bold")).pack(anchor="w")
        
        # Stock concentration
        if factor_name != "buffer pH":
            stock_frame = ttk.LabelFrame(info_frame, text="Stock Concentration (required)", padding=8)
            stock_frame.pack(fill=tk.X, pady=(8, 0))
            
            # Concentration entry and unit dropdown
            entry_frame = ttk.Frame(stock_frame)
            entry_frame.pack(fill=tk.X)
            
            ttk.Label(entry_frame, text="Value:").pack(side=tk.LEFT, padx=(0, 5))
            self.stock_var = tk.StringVar(value=str(stock_conc) if stock_conc else "")
            stock_entry = ttk.Entry(entry_frame, textvariable=self.stock_var, width=12,
                                   validate='key', validatecommand=vcmd_stock)
            stock_entry.pack(side=tk.LEFT, padx=(0, 10))
            stock_entry.bind('<Return>', lambda e: self._try_save())  # Enter key to proceed
            self.stock_entry_widget = stock_entry  # Store for key binding
            
            ttk.Label(entry_frame, text="Unit:").pack(side=tk.LEFT, padx=(0, 5))
            
            # Determine unit options based on factor name
            unit_options = self._get_unit_options(factor_name)
            self.unit_var = tk.StringVar(value=unit_options[0])
            unit_dropdown = ttk.Combobox(entry_frame, textvariable=self.unit_var, 
                                        values=unit_options, state="readonly", width=8)
            unit_dropdown.pack(side=tk.LEFT)
        
        # Levels editor
        levels_frame = ttk.LabelFrame(main_frame, text="Levels", padding=8)
        levels_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
        
        # Hints
        hint_label = ttk.Label(levels_frame, 
                              text="Tip: Use commas to add multiple values at once (e.g., 2, 4, 6, 8)",
                              foreground="gray", font=("TkDefaultFont", 9, "italic"))
        hint_label.pack(anchor="w", pady=(0, 5))
        
        # Frame
        quick_frame = ttk.Frame(levels_frame)
        quick_frame.pack(fill=tk.X, pady=(0, 8))
        
        ttk.Label(quick_frame, text="Add level:").pack(side=tk.LEFT)
        self.level_var = tk.StringVar()
        level_entry = ttk.Entry(quick_frame, textvariable=self.level_var, width=20,
                               validate='key', validatecommand=vcmd)
        level_entry.pack(side=tk.LEFT, padx=5)
        level_entry.bind('<Return>', lambda e: self._on_level_entry_enter())
        self.level_entry_widget = level_entry  # Store for key binding
        ttk.Button(quick_frame, text="Add", command=self._add_level).pack(side=tk.LEFT)
        
        # Levels listbox
        list_frame = ttk.Frame(levels_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        self.levels_listbox = tk.Listbox(list_frame, height=6, exportselection=False)
        self.levels_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Bind Backspace and Delete keys to remove selected item
        self.levels_listbox.bind('<BackSpace>', lambda e: self._remove_level())
        self.levels_listbox.bind('<Delete>', lambda e: self._remove_level())
        
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", 
                                 command=self.levels_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.levels_listbox.configure(yscrollcommand=scrollbar.set)
        
        # Populate existing levels
        if existing_levels:
            for level in existing_levels:
                self.levels_listbox.insert(tk.END, level)
        
        # Level controls
        ctrl_frame = ttk.Frame(levels_frame)
        ctrl_frame.pack(fill=tk.X, pady=(8, 0))
        ttk.Button(ctrl_frame, text="Remove Selected", 
                  command=self._remove_level).pack(side=tk.LEFT, padx=2)
        ttk.Button(ctrl_frame, text="Clear All", 
                  command=self._clear_levels).pack(side=tk.LEFT, padx=2)
        
        # Action buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X)
        
        ttk.Button(btn_frame, text="Save", command=self._save).pack(side=tk.RIGHT, padx=2)
        ttk.Button(btn_frame, text="Cancel", command=self.destroy).pack(side=tk.RIGHT, padx=2)
        
        # Bind Enter key to save
        def on_enter(event):

            if event.widget == self.level_entry_widget or event.widget == self.stock_entry_widget:
                return
            self._save()
        
        self.bind('<Return>', on_enter)
    
    def _on_level_entry_enter(self):
        """Smart Enter key behavior for level entry:
        - If entry has text: add it as a level
        - If entry is empty and levels exist: save the dialog
        """
        level_text = self.level_var.get().strip()
        
        if level_text:

            self._add_level()
        else:

            if self.levels_listbox.size() > 0:
                self._save()
    
    def _add_level(self):
        level_input = self.level_var.get().strip()
        if level_input:

            if ',' in level_input:
                # Split by comma and add each level separately
                new_levels = [l.strip() for l in level_input.split(',') if l.strip()]
                existing = list(self.levels_listbox.get(0, tk.END))
                
                for level in new_levels:
                    # Parse to extract just the number (removes units like "mM", "%", etc.)
                    parsed_value = self._parse_numeric_value(level)
                    if parsed_value is not None:
                        level_clean = str(parsed_value)
                    else:
                        level_clean = level
                    
                    if level_clean not in existing:
                        self.levels_listbox.insert(tk.END, level_clean)
                        existing.append(level_clean)
                
                self.level_var.set("")
            else:

                parsed_value = self._parse_numeric_value(level_input)
                if parsed_value is not None:
                    level_clean = str(parsed_value)
                else:
                    level_clean = level_input
                
                existing = self.levels_listbox.get(0, tk.END)
                if level_clean not in existing:
                    self.levels_listbox.insert(tk.END, level_clean)
                    self.level_var.set("")
                else:
                    messagebox.showwarning("Duplicate", f"Level '{level_clean}' already exists.")
    
    def _remove_level(self):
        selection = self.levels_listbox.curselection()
        if selection:
            self.levels_listbox.delete(selection[0])
    
    def _clear_levels(self):
        if messagebox.askyesno("Confirm", "Clear all levels?"):
            self.levels_listbox.delete(0, tk.END)
    
    def _sort_levels(self):
        items = list(self.levels_listbox.get(0, tk.END))
        try:
            # Try numeric sort
            sorted_items = sorted(items, key=lambda x: float(x))
        except ValueError:

            sorted_items = sorted(items)
        
        self.levels_listbox.delete(0, tk.END)
        for item in sorted_items:
            self.levels_listbox.insert(tk.END, item)
    
    def _get_unit_options(self, factor_name: str) -> List[str]:
        """Get appropriate unit options based on factor name"""
        factor_lower = factor_name.lower()
        

        if any(x in factor_lower for x in ['%', 'percent', 'glycerol', 'dmso', 'detergent', 'tween']):
            return ['w/v', 'v/v']
        

        if 'ph' in factor_lower:
            return ['pH units']
        

        return ['mM', 'µM', 'M', 'nM']
    
    def _convert_to_base_unit(self, value: float, unit: str, factor_name: str) -> float:
        """Convert value to base unit (mM for concentrations, % for percentages)"""
        factor_lower = factor_name.lower()
        

        if any(x in factor_lower for x in ['%', 'percent', 'glycerol', 'dmso', 'detergent', 'tween']):
            if unit in ['w/v', 'v/v']:
                return value
            return value
        

        if 'ph' in factor_lower:
            return value
        
        # For concentration factors, convert everything to mM
        conversions = {
            'M': 1000,         # 1 M = 1000 mM
            'mM': 1,           # base unit
            'µM': 0.001,       # 1 µM = 0.001 mM
            'nM': 0.000001,    # 1 nM = 0.000001 mM
        }
        
        return value * conversions.get(unit, 1)
    
    def _validate_range(self, value: float, factor_name: str) -> Tuple[bool, str]:
        """Validate that value is within acceptable range for the factor type"""
        factor_lower = factor_name.lower()
        
        # pH validation: 0-14
        if 'ph' in factor_lower and 'buffer' not in factor_lower:
            if value < 0 or value > 14:
                return False, f"pH must be between 0 and 14.\n\nYou entered: {value}\n\nPlease enter a realistic pH value."
        
        # Percentage validation: 0-100
        if any(x in factor_lower for x in ['%', 'percent', 'glycerol', 'dmso', 'detergent', 'tween']):
            if value < 0 or value > 100:
                return False, f"Percentage must be between 0 and 100%. You entered: {value}%"
        
        # General validation: no negative concentrations
        if value < 0:
            return False, f"Concentration cannot be negative. You entered: {value}"
        
        return True, ""
    
    def _parse_numeric_value(self, text: str) -> Optional[float]:
        """Extract numeric value from text, ignoring units and other characters"""

        cleaned = re.sub(r'[a-zA-Z/%°µ\s]+', '', text)
        try:
            return float(cleaned)
        except ValueError:
            return None
    
    def _try_save(self):
        """Try to save - called by Enter key. Only saves if we have levels."""
        levels = list(self.levels_listbox.get(0, tk.END))
        if levels:

            self._save()

    
    def _save(self):
        levels = list(self.levels_listbox.get(0, tk.END))
        if not levels:
            messagebox.showerror("Error", "At least one level is required.")
            return
        
        # MANDATORY stock concentration for all except buffer pH
        stock = None
        if hasattr(self, 'stock_var'):
            stock_str = self.stock_var.get().strip()
            if not stock_str:
                messagebox.showerror("Stock Concentration Required", 
                    "You must enter a stock concentration.\n\n"
                    "This is required to calculate volumes for the Opentrons.")
                return
            
            # Parse numeric value
            stock_value = self._parse_numeric_value(stock_str)
            if stock_value is None:
                try:
                    stock_value = float(stock_str)
                except ValueError:
                    messagebox.showerror("Error", 
                        f"Invalid stock concentration: '{stock_str}'\n\n"
                        f"Please enter a numeric value.")
                    return
            
            if stock_value <= 0:
                messagebox.showerror("Error", "Stock concentration must be positive.")
                return
            
            # Get selected unit and convert to base unit
            unit = self.unit_var.get()
            stock_base = self._convert_to_base_unit(stock_value, unit, self.factor_name)
            
            # Validate range
            is_valid, error_msg = self._validate_range(stock_base, self.factor_name)
            if not is_valid:
                messagebox.showerror("Invalid Range", error_msg)
                return
            
            stock = stock_base
            
            # Check that no level exceeds stock concentration
            invalid_levels = []
            for level in levels:
                level_value = self._parse_numeric_value(level)
                if level_value is None:
                    try:
                        level_value = float(level)
                    except ValueError:
                        continue
                
                is_valid, error_msg = self._validate_range(level_value, self.factor_name)
                if not is_valid:
                    messagebox.showerror("Invalid Level Range", error_msg)
                    return
                
                # Check if exceeds stock
                if level_value > stock_base:
                    invalid_levels.append(f"{level} (parsed as {level_value})")
            
            if invalid_levels:
                messagebox.showerror("Invalid Concentrations", 
                    f"The following concentration(s) EXCEED the stock concentration:\n\n"
                    f"Stock: {stock_base} {unit}\n"
                    f"Invalid levels: {', '.join(invalid_levels)}\n\n"
                    f"⚠️ You cannot make a solution more concentrated than your stock!\n\n"
                    f"Either:\n"
                    f"• Increase stock concentration, OR\n"
                    f"• Reduce the level values")
                return
        
        # Sort levels numerically if possible
        try:
            levels = sorted(levels, key=lambda x: float(x))
        except ValueError:
            # fallback to string sort if not all numeric
            levels = sorted(levels)
        
        self.result_levels = levels
        self.result_stock = stock
        self.destroy()


class FactorialDesigner(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Factorial Designer v0.9.4")
        self.geometry("1000x600")
        self.minsize(900, 550)
        
        self.model = FactorModel()
        self._build_ui()
        self._update_display()
        
        self.protocol("WM_DELETE_WINDOW", self._on_closing)
    
    def _build_ui(self):
        # Configure grid
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        
        # Main container with reduced padding
        main_container = ttk.Frame(self, padding=8)
        main_container.grid(row=0, column=0, sticky="nsew")
        main_container.columnconfigure(0, weight=1)
        main_container.columnconfigure(1, weight=2)
        main_container.rowconfigure(0, weight=1)
        main_container.rowconfigure(1, weight=0)
        
        # Left panel: Factor Selection
        left_panel = ttk.Frame(main_container)
        left_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 4))
        left_panel.columnconfigure(0, weight=1)
        left_panel.rowconfigure(0, weight=0)
        left_panel.rowconfigure(1, weight=1)
        
        # Quick actions: Custom factor
        quick_frame = ttk.LabelFrame(left_panel, text="Quick Actions", padding=8)
        quick_frame.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        
        ttk.Button(quick_frame, text="Add Custom Factor",
                  command=self._add_custom_factor).pack(fill=tk.X, pady=2)
        ttk.Button(quick_frame, text="Clear All",
                  command=self._clear_all).pack(fill=tk.X, pady=2)
        
        # Available factors
        factors_frame = ttk.LabelFrame(left_panel, text="Available Factors", padding=8)
        factors_frame.grid(row=1, column=0, sticky="nsew")
        factors_frame.columnconfigure(0, weight=1)
        factors_frame.rowconfigure(1, weight=1)
        
        ttk.Label(factors_frame, text="Double-click to add:", 
                 font=("TkDefaultFont", 9, "italic")).grid(row=0, column=0, sticky="w", pady=(0, 4))
        
        self.available_listbox = tk.Listbox(factors_frame, exportselection=False)
        self.available_listbox.grid(row=1, column=0, sticky="nsew")
        self.available_listbox.bind('<Double-Button-1>', lambda e: self._quick_add_factor())
        
        scrollbar = ttk.Scrollbar(factors_frame, orient="vertical", 
                                 command=self.available_listbox.yview)
        scrollbar.grid(row=1, column=1, sticky="ns")
        self.available_listbox.configure(yscrollcommand=scrollbar.set)
        
        # Populate available factors
        for key, display in AVAILABLE_FACTORS.items():
            self.available_listbox.insert(tk.END, display)
        
        # Right panel: Current Design
        right_panel = ttk.Frame(main_container)
        right_panel.grid(row=0, column=1, sticky="nsew", padx=(4, 0))
        right_panel.columnconfigure(0, weight=1)
        right_panel.rowconfigure(0, weight=1)
        
        # Design table
        design_frame = ttk.LabelFrame(right_panel, text="Current Design", padding=8)
        design_frame.grid(row=0, column=0, sticky="nsew")
        design_frame.columnconfigure(0, weight=1)
        design_frame.rowconfigure(0, weight=1)
        
        # Treeview for factors
        self.tree = ttk.Treeview(design_frame, 
                                columns=("factor", "levels", "count", "stock"),
                                show="headings", height=12, selectmode="browse")
        self.tree.heading("factor", text="Factor")
        self.tree.heading("levels", text="Levels")
        self.tree.heading("count", text="# Levels")
        self.tree.heading("stock", text="Stock Conc")
        
        self.tree.column("factor", width=130, anchor="w")
        self.tree.column("levels", width=250, anchor="w")
        self.tree.column("count", width=70, anchor="center")
        self.tree.column("stock", width=90, anchor="center")
        
        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.bind('<Double-Button-1>', lambda e: self._edit_factor())
        
        tree_scroll = ttk.Scrollbar(design_frame, orient="vertical", command=self.tree.yview)
        tree_scroll.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=tree_scroll.set)
        
        # Instructions
        ttk.Label(design_frame, text="Double-click a factor to edit",
                 font=("TkDefaultFont", 8, "italic")).grid(row=1, column=0, sticky="w", pady=(4, 0))
        
        # Factor controls
        ctrl_frame = ttk.Frame(design_frame)
        ctrl_frame.grid(row=2, column=0, sticky="ew", pady=(6, 0))
        
        ttk.Button(ctrl_frame, text="Edit",
                  command=self._edit_factor).pack(side=tk.LEFT, padx=2)
        ttk.Button(ctrl_frame, text="Delete",
                  command=self._delete_factor).pack(side=tk.LEFT, padx=2)
        
        # Bottom panel: Export Settings
        bottom_panel = ttk.LabelFrame(main_container, text="Export", padding=8)
        bottom_panel.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(8, 0))
        bottom_panel.columnconfigure(1, weight=1)
        
        # Final volume
        ttk.Label(bottom_panel, text="Final Volume (µL):").grid(row=0, column=0, sticky="w")
        self.final_vol_var = tk.StringVar(value="100")
        vcmd = (self.register(validate_single_numeric_input), '%d', '%S', '%P')
        ttk.Entry(bottom_panel, textvariable=self.final_vol_var, width=12,
                 validate='key', validatecommand=vcmd).grid(row=0, column=1, sticky="w", padx=5)
        
        # Status display
        status_frame = ttk.Frame(bottom_panel)
        status_frame.grid(row=0, column=2, sticky="e")
        
        ttk.Label(status_frame, text="Combinations:").pack(side=tk.LEFT, padx=(0, 3))
        self.combo_var = tk.StringVar(value="0")
        combo_label = ttk.Label(status_frame, textvariable=self.combo_var, 
                               font=("TkDefaultFont", 10, "bold"), foreground="#2196F3")
        combo_label.pack(side=tk.LEFT, padx=(0, 8))
        
        ttk.Label(status_frame, text="Plates:").pack(side=tk.LEFT, padx=(0, 3))
        self.plates_var = tk.StringVar(value="0")
        plates_label = ttk.Label(status_frame, textvariable=self.plates_var, 
                                font=("TkDefaultFont", 10, "bold"), foreground="#4CAF50")
        plates_label.pack(side=tk.LEFT)
        
        # Export button
        ttk.Button(bottom_panel, text="Export Design",
                  command=self._export_both).grid(row=1, column=0, columnspan=3, pady=(8, 0))
    
    def _add_custom_factor(self):
        """Add a completely custom factor (e.g., KCl, MgCl2, etc.) with units"""
        # Ask for custom factor name with units
        dialog = tk.Toplevel(self)
        dialog.title("Add Custom Factor")
        dialog.geometry("450x250")
        dialog.transient(self)
        dialog.grab_set()
        
        result = {"name": None}
        
        frame = ttk.Frame(dialog, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Enter custom factor name:", 
                 font=("TkDefaultFont", 10, "bold")).pack(anchor="w", pady=(0, 10))
        
        ttk.Label(frame, text="Examples: KCl, MgCl2, DTT, EDTA, Tween-20").pack(anchor="w", pady=(0, 5))
        
        ttk.Label(frame, text="Note: Units will be selected in the next step", 
                 foreground="gray", font=("TkDefaultFont", 9, "italic")).pack(anchor="w", pady=(0, 10))
        
        name_var = tk.StringVar()
        name_entry = ttk.Entry(frame, textvariable=name_var, width=35)
        name_entry.pack(fill=tk.X, pady=(0, 15))
        name_entry.focus_set()
        
        def on_ok():
            name = name_var.get().strip()
            if not name:
                messagebox.showwarning("Invalid Name", "Please enter a factor name.")
                return
            
            result["name"] = name
            dialog.destroy()
        
        def on_cancel():
            dialog.destroy()
        
        name_entry.bind('<Return>', lambda e: on_ok())
        
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=tk.X)
        ttk.Button(btn_frame, text="OK", command=on_ok).pack(side=tk.RIGHT, padx=2)
        ttk.Button(btn_frame, text="Cancel", command=on_cancel).pack(side=tk.RIGHT, padx=2)
        
        self.wait_window(dialog)
        
        if result["name"]:
            custom_name = result["name"]
            

            factors = self.model.get_factors()
            if custom_name in factors:
                messagebox.showwarning("Duplicate", f"Factor '{custom_name}' already exists.")
                return
            

            editor_dialog = FactorEditDialog(self, custom_name)
            self.wait_window(editor_dialog)
            
            if editor_dialog.result_levels:
                self.model.add_factor(custom_name, editor_dialog.result_levels, editor_dialog.result_stock)
                self._update_display()
    
    def _add_new_factor(self):
        """Show factor selection and add dialog (from predefined list)"""
        selection = self.available_listbox.curselection()
        if not selection:
            messagebox.showinfo("Select Factor", "Please select a factor from the list or double-click it.")
            return
        
        idx = selection[0]
        factor_key = list(AVAILABLE_FACTORS.keys())[idx]
        

        if factor_key in self.model.get_factors():
            messagebox.showwarning("Duplicate", f"Factor '{AVAILABLE_FACTORS[factor_key]}' already exists.")
            return
        

        dialog = FactorEditDialog(self, factor_key)
        self.wait_window(dialog)
        
        if dialog.result_levels:
            self.model.add_factor(factor_key, dialog.result_levels, dialog.result_stock)
            self._update_display()
    
    def _quick_add_factor(self):
        """Quick add from double-click"""
        self._add_new_factor()
    
    def _edit_factor(self):
        """Edit selected factor"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showinfo("Select Factor", "Please select a factor to edit.")
            return
        
        item = selection[0]
        factor_display = self.tree.item(item)['values'][0]
        

        internal_key = None
        for key in self.model.get_factors().keys():
            if AVAILABLE_FACTORS.get(key, key) == factor_display:
                internal_key = key
                break
        
        if not internal_key:
            return
        

        existing_levels = self.model.get_factors()[internal_key]
        existing_stock = self.model.get_stock_conc(internal_key)
        

        dialog = FactorEditDialog(self, internal_key, existing_levels, existing_stock)
        self.wait_window(dialog)
        
        if dialog.result_levels:
            self.model.update_factor(internal_key, dialog.result_levels, dialog.result_stock)
            self._update_display()
    
    def _delete_factor(self):
        """Delete selected factor"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showinfo("Select Factor", "Please select a factor to delete.")
            return
        
        item = selection[0]
        factor_display = self.tree.item(item)['values'][0]
        
        if messagebox.askyesno("Confirm", f"Delete factor '{factor_display}'?"):

            for key in self.model.get_factors().keys():
                if AVAILABLE_FACTORS.get(key, key) == factor_display:
                    self.model.remove_factor(key)
                    break
            self._update_display()
    
    def _clear_all(self):
        """Clear all factors"""
        if messagebox.askyesno("Confirm", "Clear all factors?"):
            self.model.clear()
            self._update_display()
    
    def _update_display(self):
        """Update the factors table and stats - REAL-TIME"""

        for item in self.tree.get_children():
            self.tree.delete(item)
        

        factors = self.model.get_factors()
        for key, levels in factors.items():
            display_name = AVAILABLE_FACTORS.get(key, key)
            levels_preview = ", ".join(levels[:5])
            if len(levels) > 5:
                levels_preview += "..."
            
            stock = self.model.get_stock_conc(key)
            stock_display = f"{stock}" if stock else "—"
            

            num_levels = len(levels)
            
            self.tree.insert("", tk.END, values=(display_name, levels_preview, 
                                                num_levels, stock_display))
        

        total = self.model.total_combinations()
        self.combo_var.set(str(total))
        
        plates = (total + 95) // 96 if total > 0 else 0
        self.plates_var.set(str(plates))
        
        # Warn if too many
        if total > 384:
            self.combo_var.set(f"{total} ⚠️")
            self.plates_var.set(f"{plates} ⚠️")
        
        self.update_idletasks()
    
    def _generate_well_position(self, index: int) -> Tuple[int, str]:
        """Generate plate number and well position (A1-H12)
        Fills column-wise: A1, B1, C1...H1, A2, B2...
        """
        plate_number = (index // 96) + 1
        position_in_plate = index % 96
        
        # Fill down columns first (A1, B1, C1... H1, then A2, B2...)
        row = position_in_plate % 8  # 0-7 for A-H
        col = position_in_plate // 8  # 0-11 for 1-12
        
        row_letter = chr(ord('A') + row)
        col_number = col + 1
        well = f"{row_letter}{col_number}"
        
        return plate_number, well
    
    def _convert_96_to_384_well(self, plate_num: int, well_96: str) -> str:
        """Convert 96-well to 384-well position
        
        Mapping rules:
        - Plate 1→ cols 1-6, Plate 2→ cols 7-12, etc.
        - Odd cols (1,3,5) use first row of pair (A→A, B→C, C→E)
        - Even cols (2,4,6) use second row (A→B, B→D, C→F)
        """
        import math
        
        # Parse 96-well position (e.g., "B3" → row=B, col=3)
        row_96 = well_96[0]  # Letter (A-H)
        col_96 = int(well_96[1:])  # Number (1-12)
        
        # Convert row letter to index (A=0, B=1, ..., H=7)
        row_96_index = ord(row_96) - ord('A')
        
        # Map to 384-well column: (plate-1)*6 + ceil(col/2)
        col_384 = (plate_num - 1) * 6 + math.ceil(col_96 / 2)
        
        # Map 96-well row to 384-well row based on column parity
        # Each 96 row → 2 consecutive 384 rows (odd col→first, even col→second)
        if col_96 % 2 == 1:  # Odd column
            row_384_index = row_96_index * 2
        else:  # Even column
            row_384_index = row_96_index * 2 + 1
        
        # Convert back to letter (A=0, B=1, ..., P=15)
        row_384 = chr(ord('A') + row_384_index)
        
        return f"{row_384}{col_384}"
    
    def _build_factorial_with_volumes(self) -> Tuple[List[str], List[List], List[str], List[List]]:
        """Build full factorial design and calculate volumes"""
        factors = self.model.get_factors()
        if not factors:
            raise ValueError("No factors defined.")
        
        # Get final volume
        try:
            final_vol = float(self.final_vol_var.get())
        except ValueError:
            raise ValueError("Invalid final volume value.")
        
        stock_concs = self.model.get_all_stock_concs()
        
        # Build factorial combinations
        factor_names = list(factors.keys())
        level_lists = [factors[f] for f in factor_names]
        combinations = list(itertools.product(*level_lists))
        
        # Determine buffer pH values - make unique pH list for Opentrons
        buffer_ph_values = []
        if "buffer pH" in factors:
            # Extract unique pH values from the levels
            unique_phs = set()
            for ph_val in factors["buffer pH"]:
                unique_phs.add(str(ph_val).strip())
            buffer_ph_values = sorted(unique_phs)
        
        # Build Excel headers
        excel_headers = ["ID", "Plate_96", "Well_96", "Well_384"]
        for fn in factor_names:
            excel_headers.append(AVAILABLE_FACTORS.get(fn, fn))
        # Add Response column for TM data entry
        excel_headers.append("Response")
        
        # Build volume headers
        volume_headers = []
        if "buffer pH" in factors:
            for ph in buffer_ph_values:
                volume_headers.append(f"buffer {ph}")
        
        # Add other volume headers

        for factor in factor_names:
            if factor in ["buffer pH", "buffer_concentration"]:
                continue
            volume_headers.append(factor)
        
        # Add water column at the end
        volume_headers.append("water")
        
        # Calculate volumes
        excel_rows = []
        volume_rows = []
        
        for idx, combo in enumerate(combinations):
            row_dict = {factor_names[i]: combo[i] for i in range(len(factor_names))}
            
            plate_num, well_pos = self._generate_well_position(idx)
            well_384 = self._convert_96_to_384_well(plate_num, well_pos)
            
            # Excel row
            excel_row = [idx + 1, plate_num, well_pos, well_384]
            for fn in factor_names:
                excel_row.append(row_dict.get(fn, ""))
            # Add empty Response column for manual data entry
            excel_row.append("")
            excel_rows.append(excel_row)
            
            # Volume calculations
            volumes = {}
            total_volume_used = 0
            
            # Handle buffer pH
            if "buffer pH" in row_dict:
                buffer_ph = str(row_dict["buffer pH"])
                for ph in buffer_ph_values:
                    volumes[f"buffer {ph}"] = 0
                
                if "buffer_concentration" in row_dict and "buffer_concentration" in stock_concs:
                    try:
                        desired_conc = float(row_dict["buffer_concentration"])
                        buffer_stock = stock_concs["buffer_concentration"]
                        # C1*V1 = C2*V2 → V1 = (C2*V2)/C1
                        volume = (desired_conc * final_vol) / buffer_stock
                        volumes[f"buffer {buffer_ph}"] = round(volume, 2)
                        total_volume_used += volumes[f"buffer {buffer_ph}"]
                    except (ValueError, ZeroDivisionError):
                        volumes[f"buffer {buffer_ph}"] = 0
            
            # Calculate volumes for other factors (glycerol, salt, DMSO, etc.)
            for factor in factor_names:
                if factor in ["buffer pH", "buffer_concentration"]:
                    continue
                if factor in row_dict and factor in stock_concs:
                    try:
                        desired_conc = float(row_dict[factor])
                        stock_conc = stock_concs[factor]
                        volume = (desired_conc * final_vol) / stock_conc
                        volumes[factor] = round(volume, 2)
                        total_volume_used += volumes[factor]
                    except (ValueError, ZeroDivisionError):
                        volumes[factor] = 0
            
            # Calculate water to reach final volume
            water_volume = round(final_vol - total_volume_used, 2)
            volumes["water"] = water_volume
            
            volume_row = [volumes.get(h, 0) for h in volume_headers]
            volume_rows.append(volume_row)
        
        # Check for negative water volumes
        negative_water_wells = []
        for idx, volume_row in enumerate(volume_rows):
            water_idx = volume_headers.index("water")
            water_vol = volume_row[water_idx]
            if water_vol < 0:
                well_id = excel_rows[idx][0]  # ID
                well_pos = excel_rows[idx][2]  # Well position
                negative_water_wells.append((well_id, well_pos, water_vol))
        
        if negative_water_wells:
            # Build error message for impossible designs
            error_msg = "⚠️ IMPOSSIBLE DESIGN DETECTED ⚠️\n\n"
            error_msg += f"The following wells require NEGATIVE water volumes:\n\n"
            
            # Show problematic wells
            for well_id, well_pos, water_vol in negative_water_wells[:5]:
                error_msg += f"  • Well {well_pos} (ID {well_id}): {water_vol} µL water\n"
            
            if len(negative_water_wells) > 5:
                error_msg += f"  ... and {len(negative_water_wells) - 5} more wells\n"
            
            error_msg += f"\nTotal problematic wells: {len(negative_water_wells)}\n\n"
            error_msg += "This means the sum of component volumes EXCEEDS the final volume!\n\n"
            error_msg += "Solutions:\n"
            error_msg += "  1. INCREASE stock concentrations (recommended)\n"
            error_msg += "  2. INCREASE final volume\n"
            error_msg += "  3. REDUCE desired concentration levels\n\n"
            error_msg += "Example: If stock is 50 mM and you want 100 mM,\n"
            error_msg += "you'd need to add 200 µL of stock to make 100 µL final volume.\n"
            error_msg += "This is physically impossible!"
            
            raise ValueError(error_msg)
        
        return excel_headers, excel_rows, volume_headers, volume_rows
    
    def _export_both(self):
        """Export XLSX and CSV files with single-step file dialog"""
        if not HAS_OPENPYXL:
            messagebox.showerror("Missing Library", 
                               "openpyxl is required. Install with: pip install openpyxl")
            return
        
        try:
            # Validate stock concentrations
            factors = self.model.get_factors()
            stock_concs = self.model.get_all_stock_concs()
            
            missing_stocks = []
            for factor in factors.keys():
                if factor == "buffer pH":
                    continue
                if factor not in stock_concs:
                    missing_stocks.append(AVAILABLE_FACTORS.get(factor, factor))
            
            if missing_stocks:
                messagebox.showerror("Missing Stock Concentrations",
                    f"The following factors need stock concentrations:\n\n" +
                    "\n".join(f"• {f}" for f in missing_stocks) +
                    "\n\nEdit each factor to add stock concentrations.")
                return
            
            # Build design
            excel_headers, excel_rows, volume_headers, volume_rows = self._build_factorial_with_volumes()
            
            total = len(excel_rows)
            if total > 384:
                messagebox.showerror("Too Many Combinations",
                    f"Design has {total} combinations.\n\n"
                    f"Maximum: 384 (4 plates of 96 wells)\n\n"
                    f"Please reduce factors or levels.")
                return
            
            # Single-step file save dialog
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save Factorial Design"
            )
            
            if not path:
                return  # User cancelled
            
            # Generate paths
            base_path = os.path.splitext(path)[0]
            xlsx_path = base_path + ".xlsx"
            csv_path = base_path + "_opentron.csv"
            
            # Export XLSX
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sample Tracking"
            
            # Headers
            for col_idx, header in enumerate(excel_headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            
            # Data
            for row_idx, row_data in enumerate(excel_rows, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust columns
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass  # TODO: maybe log this instead of silent fail?
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width
            
            wb.save(xlsx_path)
            
            # Export CSV
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(volume_headers)
                for vol_row in volume_rows:
                    writer.writerow(vol_row)
            
            plates = (total + 95) // 96
            
            messagebox.showinfo("Export Successful!",
                f"Files exported successfully!\n\n"
                f"📊 Sample tracking:\n{xlsx_path}\n\n"
                f"🤖 Opentrons volumes:\n{csv_path}\n\n"
                f"Total: {total} combinations ({plates} plates)")
        
        except Exception as e:
            messagebox.showerror("Export Failed", f"Error during export:\n\n{str(e)}")
    
    def _on_closing(self):
        """Handle window closing"""
        self.destroy()


if __name__ == "__main__":
    app = FactorialDesigner()
    app.mainloop()
